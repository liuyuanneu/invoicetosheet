import os, io, json, tempfile, re
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import anthropic
from dotenv import load_dotenv

load_dotenv()
app = FastAPI(title="InvoiceToSheet")

def extract_pdf_text(pdf_bytes: bytes) -> str:
    text_parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)
    return "\n".join(text_parts)

SYSTEM_PROMPT = """You are an invoice data extraction specialist.
Extract structured data from the invoice text provided.
Return ONLY a valid JSON object — no markdown, no explanation — with this exact schema:

{
  "vendor_name": "string or null",
  "vendor_address": "string or null",
  "invoice_number": "string or null",
  "invoice_date": "string or null",
  "due_date": "string or null",
  "currency": "string (e.g. USD, EUR, GBP) or null",
  "subtotal": "number or null",
  "tax": "number or null",
  "total": "number or null",
  "line_items": [
    {
      "description": "string",
      "quantity": "number or null",
      "unit_price": "number or null",
      "amount": "number or null"
    }
  ],
  "notes": "string or null"
}

Rules:
- Parse all numbers as numeric values, not strings
- Dates in ISO format (YYYY-MM-DD) if possible, otherwise as found
- line_items must always be a list (empty list if none found)
- If a field is missing, use null
"""

def extract_invoice_data(text: str) -> dict:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="API key not configured")
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2048,
        messages=[{"role": "user", "content": f"Extract invoice data from this text:\n\n{text}"}],
        system=SYSTEM_PROMPT,
    )
    raw = message.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)

def build_excel(invoices: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1A1A2E")
    section_fill = PatternFill("solid", fgColor="EEF2FF")
    normal_font  = Font(name="Arial", size=9)
    bold_font    = Font(name="Arial", bold=True, size=9)

    def hcell(ws, row, col, value, fill=None, font=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        if fill: c.fill = fill
        if font: c.font = font
        else:    c.font = normal_font
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return c

    for idx, inv in enumerate(invoices):
        sheet_name = f"Invoice {idx+1}" if len(invoices) > 1 else "Invoice"
        ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        r = 1
        ws.merge_cells(f"A{r}:E{r}")
        title = ws.cell(row=r, column=1, value="INVOICE DATA — InvoiceToSheet.com")
        title.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        title.fill = header_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 28
        r += 1
        fields = [
            ("Vendor / Supplier", inv.get("vendor_name")),
            ("Vendor Address",    inv.get("vendor_address")),
            ("Invoice Number",    inv.get("invoice_number")),
            ("Invoice Date",      inv.get("invoice_date")),
            ("Due Date",          inv.get("due_date")),
            ("Currency",          inv.get("currency")),
        ]
        for label, value in fields:
            ws.merge_cells(f"B{r}:E{r}")
            hcell(ws, r, 1, label, fill=section_fill, font=bold_font)
            hcell(ws, r, 2, value if value else "—")
            ws.row_dimensions[r].height = 18
            r += 1
        r += 1
        ws.row_dimensions[r].height = 20
        for col, h in enumerate(["Description", "Qty", "Unit Price", "Amount"], start=2):
            hcell(ws, r, col, h, fill=PatternFill("solid", fgColor="0F3460"), font=header_font, align="center")
        r += 1
        items = inv.get("line_items") or []
        for item in items:
            ws.row_dimensions[r].height = 18
            hcell(ws, r, 2, item.get("description") or "—")
            hcell(ws, r, 3, item.get("quantity"), align="center")
            hcell(ws, r, 4, item.get("unit_price"), align="right")
            hcell(ws, r, 5, item.get("amount"), align="right")
            for col in [4, 5]:
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            r += 1
        if not items:
            ws.merge_cells(f"B{r}:E{r}")
            hcell(ws, r, 2, "No line items found")
            r += 1
        r += 1
        for label, value in [("Subtotal", inv.get("subtotal")), ("Tax", inv.get("tax")), ("TOTAL", inv.get("total"))]:
            is_total = label == "TOTAL"
            ws.merge_cells(f"B{r}:D{r}")
            hcell(ws, r, 2, label, fill=PatternFill("solid", fgColor="1A1A2E") if is_total else section_fill,
                  font=Font(name="Arial", bold=True, size=9, color="FFFFFF" if is_total else "000000"), align="right")
            c = ws.cell(row=r, column=5, value=value)
            c.border = border
            c.font = Font(name="Arial", bold=is_total, size=9, color="FFFFFF" if is_total else "000000")
            c.fill = PatternFill("solid", fgColor="1A1A2E") if is_total else section_fill
            c.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(value, (int, float)):
                c.number_format = '#,##0.00'
            ws.row_dimensions[r].height = 20
            r += 1
        if inv.get("notes"):
            r += 1
            hcell(ws, r, 1, "Notes", fill=section_fill, font=bold_font)
            ws.merge_cells(f"B{r}:E{r}")
            hcell(ws, r, 2, inv["notes"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

@app.post("/extract")
async def extract(files: list[UploadFile] = File(...)):
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 files per request")
    results = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{f.filename} is not a PDF")
        pdf_bytes = await f.read()
        if len(pdf_bytes) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"{f.filename} exceeds 10MB limit")
        text = extract_pdf_text(pdf_bytes)
        if not text.strip():
            raise HTTPException(status_code=422, detail=f"{f.filename} appears to be a scanned image PDF with no extractable text")
        try:
            data = extract_invoice_data(text)
        except anthropic.BadRequestError as e:
            err = str(e)
            if "credit balance" in err or "too low" in err:
                raise HTTPException(status_code=402,
                    detail="API credits exhausted. Please contact support.")
            raise HTTPException(status_code=400, detail=f"AI extraction failed: {err}")
        except anthropic.AuthenticationError:
            raise HTTPException(status_code=500, detail="API authentication error.")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")
        data["filename"] = f.filename
        results.append(data)
    return JSONResponse(content={"invoices": results})

@app.post("/download")
async def download(files: list[UploadFile] = File(...)):
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 files per request")
    invoices = []
    filenames = []
    for f in files:
        pdf_bytes = await f.read()
        text = extract_pdf_text(pdf_bytes)
        data = extract_invoice_data(text)
        data["filename"] = f.filename
        invoices.append(data)
        filenames.append(Path(f.filename).stem)
    excel_bytes = build_excel(invoices)
    out_name = filenames[0] + "_extracted.xlsx" if len(filenames) == 1 else "invoices_extracted.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )

@app.get("/health")
def health():
    return {"status": "ok"}

static_path = Path(__file__).parent / "static"
if static_path.exists():
    app.mount("/", StaticFiles(directory=str(static_path), html=True), name="static")
